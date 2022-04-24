import psycopg2
from time import sleep
import requests
from openpyxl import Workbook,load_workbook

#kirala izle içerikler için bekletiyorsun 
wp = load_workbook('C:/Users/HP/Desktop/ingestat/metadata.xlsx')
ws = wp.active


for row in ws['A']:
    
        
    
    url = "http://10.98.225.181/api/cms-vod-core/contentlanguagemetadatas?metadataId="+'()'.format(row.value)
    headers = {

    'Authorization': 'Bearer eyJhbGciOiJIUzUxMiJ9.eyJzdWIiOiIyNyIsInJvbGVJZCI6MSwidXNlcm5hbWUiOiJPxJ91emhhbiIsImF1dGhvcml0aWVzIjpbIlJPTEVfQURNSU4iXSwic2Vzc2lvbklkIjoiMzZjZmNiYzcyOTlkNDgyNmFkMTAyMWFiOWFlMGU2NzAiLCJpYXQiOjE2MzgwMjgzMzJ9.558-J39m75cSoXWZZUAYL7PTDmuCikx8ND6aPI_YEkvvoCRd6eCpO6B_IL1HwxFyq65pxpJDvlKeWQKjqoF37w'


    }

    payload ={

    'metadataId': '()'.format(row.value)


    }
    
    print(row.value , 'content id li içeriğe ingest atıldı')
        
        
#    print(url)
#    print(i)

    response = requests.request("GET", url, headers=headers, data=payload)
    print(response.text)
