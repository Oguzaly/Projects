import requests
import psycopg2
from openpyxl import Workbook,load_workbook
import psycopg2
import smtplib
from email.message import EmailMessage



db_host = '10.98.225.186'
db_name = 'atlas_db'
db_user = 'saatcms'
db_pass = 'cms123'
port = '5000'


Email_user='oguzhan@saatteknoloji.com'
password='Oo1+1993'

conn=psycopg2.connect(dbname=db_name,user=db_user,password=db_pass,host=db_host,port=port)
cur=conn.cursor()


wp = load_workbook('C:/Users/HP/Desktop/playground/EpgConfMail/ChannelId/ChannelID.xlsx')
ws = wp.active


for a in ws['A']:


    cur.execute("select p.epg_program_id , c.name , p.release_date ,to_timestamp(p.start_time/1000)::time , to_timestamp(p.end_time/1000)::time  from atlas_cms_btv.program p left join atlas_cms_btv.program p2 on p2.epg_program_id =p.epg_program_id left join atlas_cms_btv.channel c on c.id =p.channel_id where p.release_date > (current_date-7) and c.id ="+'{}'.format(a.value)+"order by p.release_date,to_timestamp(p.start_time/1000)::time")
    frsh =cur.fetchall()
    print(len(frsh))

    try:

        with open('C:/Users/HP/Desktop/playground/EpgConfMail/Conflict.csv','a') as f:
            for i in range(0,len(frsh)):

                if frsh[i][4] == frsh[i+1][3]:
                    pass
                    print(i,frsh[i][4],frsh[i+1][3])
                    print(i)

                else:
                    print(i,frsh[i][0],frsh[i][2],frsh[i][4],frsh[i+1][3],'Conflict')
                    f.write('{}'.format(frsh[i][0])+','+'{}'.format(frsh[i][1])+','+'{}'.format(frsh[i][2])+','+'{}'.format(frsh[i][3])+','+'{}'.format(frsh[i][4])+'\n')

    except Exception as e:
        print('Bitti')
        pass


    # yaz = 'Aşağıdaki Sezonlarda altında published episode yok \n\n'+'guid'+'         '+'Name\n'+str(i[0])+' '+str(i[1])
yaz = 'Ekteki epglerin zamanları sorunludur. '
msg = EmailMessage()
msg['Subject'] = 'Time Conflict Epgler'
msg['From'] = Email_user
msg['to'] = 'Oguzhan@saatteknoloji.com'
msg.set_content(str(yaz))

csvfiles = ['C:/Users/HP/Desktop/playground/EpgConflict/Conflict.csv']

for file in csvfiles:

    with open(file, 'rb') as f:
        file_data = f.read()
        # file_type = imghdr.what(f.name)
        file_name = 'Conflict.csv'

    msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)


with smtplib.SMTP('smtp.office365.com',587) as smtp :
    smtp.ehlo()
    smtp.starttls()
    smtp.ehlo()
    smtp.login(Email_user,password)


    smtp.send_message(msg)
