from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from time import sleep
from openpyxl import Workbook,load_workbook
import pytest
driver = webdriver.Chrome('C:/Users/HP/Desktop/Oğuz/workfile/driver/chromedriver.exe')
driver.get('http://***')
driver.maximize_window()
class namegenerate:
    # Aynı ismin birden fazla kullanılmasına izin verilmediği zaman bu kullanılabilir
    def __init__(self):
            name_length = 32
            characters ='ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz1234567890'
            name =""
            for index in range(name_length):
                name  = name + random.choice(characters)
            self.name = 'Test - ' + name
class check:
    #tastermessage da(sağ alt taraftaki) 200 olup olmadığına bakar
    def __init__(self):
        self.compare=driver.find_element(by=By.XPATH,value="//div[@id='toast-container']/div/div").text
        sleep(1)
        if self.compare in '200':
            self.compare = True
        else:
            self.compare = False
            sleep(2)
            driver.quit()
        self.donecompare=driver.find_element(by=By.XPATH,value="//div[@id='toast-container']/div/div").click()
        sleep(0.5)
class checkurl:
    #Sayfa Geçişlerinde url e bakar
    def __init__(self,element,url):
        self.element=element
        self.url=url
        sleep(1)
        if self.element in self.url:
            self.compare = True
        else:
            self.compare = False
            sleep(2)
            driver.quit()
        sleep(1)
class dropdown:
    def __init__(self,element,text):
        self.element = element
        self.text = text
        sleep(0.5)
        self.element.send_keys(self.text)
        sleep(0.5)
        self.element.send_keys(Keys.DOWN,Keys.ENTER)
        sleep(0.5)
def enter():
    global TestLogin
    try:
        gelismis = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"button[id=details-button]"))).click()
        git = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"a[id='proceed-link']"))).click()
    except Exception as e:
        pass
    login = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"input[placeholder='Username']")))
    password = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"input[placeholder='Password']")))
    login.send_keys('***')
    password.send_keys('***')
    gitlogin = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"button[class='btn btn-primary btn-block']")))
    gitlogin.click()
    try:
        loginalready = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH,'//button[contains(.,"Login")]'))).click()
    except Exception as e:
        pass
    #firstelement = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"span[class='font-weight-semibold font-size-14']")))
    element1=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//span[contains(.,'ATLAS')]")))
    TestLogin=driver.current_url #Giriş yaptıktan sonra gelen dashboard sayfas
    TestLogin =checkurl(TestLogin,'***')
enter()
def dealmanagement():
    global TestDealManagement
    element1=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//span[contains(.,'ATLAS')]"))).click() #Yuhiiiiii buldummm Thank Selenium Ide
    element2=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//span[contains(.,'VOD Operations')]"))).click()
    element3=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//span[contains(.,'Deal Management')]"))).click()
    contractelement = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//a[contains(.,'Contract Management')]")))#Sayfadaki bir element
    TestDealManagement=driver.current_url
    TestDealManagement = checkurl(TestDealManagement,'***')
dealmanagement()
def licensemanagement():
    global url3
    global lisansaddmessage
    global Toasteradd
    wp = load_workbook('C:/Users/HP/Desktop/playground/Regresyon/DummyData/licensedata.xlsx')
    ws = wp.active

    licenseelement = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//a[contains(.,'License Management')]"))).click()
    NewLicense = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//button[contains(.,'New License ')]"))).click()

    for i in range(2,3):
        licensename = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='License Name']"))).send_keys(ws['A{}'.format(i)].value)
        ServiceType = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Select Type']"))).send_keys(ws['B{}'.format(i)].value)
        LicenseStartDate = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='License Start Date']"))).send_keys(ws['C{}'.format(i)].value,Keys.TAB)
        LicenseEndDate = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='License End Date']"))).send_keys(ws['D{}'.format(i)].value,Keys.TAB)
        Contract = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Select Contract']")))
        dropdown(Contract,'Oğuzhan')
        Platform = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Select platform']")))
        platformtext=['IPTV','Web','Mobile','Smart']
        for i in platformtext:
            dropdown(Platform,'{}'.format(i))
        Region = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Select region']")))
        dropdown(Region,'ALL')
        sleep(0.5)
        Currency = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Select Currency']")))
        dropdown(Currency,'t')
        runType = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//select[@name='runType']")))
        dropdown(runType,'u')
        BoxOffice = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Select Box Office']")))
        dropdown(BoxOffice,'u')
        Add = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//button[contains(.,' Add ')]"))).click()
        sleep(0.5)
        Toasteradd =check()
    url3=driver.current_url
    url3=checkurl(url3,'***')
licensemanagement()
def licensemanagementedit():
    global url4
    global message
    global ToasterSave
    wp = load_workbook('C:/Users/HP/Desktop/playground/Regresyon/DummyData/licensedata.xlsx')
    ws = wp.active
    lisansTab=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//a[contains(.,'License Management')]"))).click()
    lisansedit=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//tr[1]/td[8]/div/div[1]/a/i"))).click()
    sleep(1)
    for i in range(3,4):
        sleep(0.5)
        name=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@name='license.name']")))
        sleep(0.5)
        name.clear()
        sleep(0.5)
        name.send_keys(ws['A{}'.format(i)].value)
        sleep(0.5)
        ServiceType=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Select type']")))
        sleep(0.5)
        ServiceType.clear()
        sleep(0.5)
        ServiceType.send_keys(ws['B{}'.format(i)].value,Keys.TAB)
        sleep(0.5)
        startdate= WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@name='startDate']")))
        sleep(0.5)
        startdate.send_keys(Keys.CONTROL+'a',Keys.DELETE)
        sleep(0.5)
        startdate.send_keys(ws['C{}'.format(i)].value,Keys.TAB)
        sleep(0.5)
        enddate= WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@name='endDate']")))
        sleep(0.5)
        enddate.send_keys(Keys.CONTROL+'a',Keys.DELETE)
        sleep(0.5)
        enddate.send_keys(ws['D{}'.format(i)].value,Keys.TAB)
        sleep(0.5)
        save=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//button[contains(.,'Save')]"))).click()
        sleep(0.5)
        ToasterSave = check()
    url4=driver.current_url
    url4=checkurl(url4,'***')
licensemanagementedit()
def licensedelete():
    global ToasterDelete
    deletelastlicense=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//tr[1]/td[8]/div/div[2]/a/i"))).click()
    sleep(0.5)
    lisansdelete = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//button[contains(.,'Delete')]"))).click()
    sleep(0.5)
    ToasterDelete = check()
licensedelete()
class TestClass:
    def test_mainboard(self):
        assert TestLogin
    def test_dealmanagement(self):
        assert TestDealManagement
    def test_addlisans(self):
        assert Toasteradd
    def test_urlcontrol3(self):
        assert url3
    def test_lisanssave(self):
        assert ToasterSave
    def test_urlcontrol4(self):
        assert url4
    def test_lisansdelete(self):
        assert ToasterSave
