import requests
import json
from openpyxl import Workbook,load_workbook


wp = load_workbook('C:/Users/HP/Desktop/playground/CreateContent/temp/createvod.xlsx')
ws = wp.active

url = "http://10.98.228.146:8090/contents"


with open ('C:/Users/HP/Desktop/playground/CreateContent/temp/createcontent.json','r+') as f :

    #
    for i in ws['A']:

    #     print (ws['A{}'.format(i)].value)

        data = json.load(f)
        data['name'] = '{}'.format(**i.value) # <--- add `id` value.
        #data['name'] ="Oguzz"
        f.seek(0)        # <--- should reset file position to the beginning.
        json.dump(data, f, indent=4)
        f.truncate()     # remove remaining part
#
# with open ('C:/Users/HP/Desktop/playground/CreateContent/temp/createcontent.json','r+') as f :
#     payload=json.load(f)
#     print(payload)
#
#
# response = requests.request("POST", url, headers=headers, data=payload)
#
# print(response.text)
