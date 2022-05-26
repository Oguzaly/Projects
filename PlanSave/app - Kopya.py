
import pandas as pd
import requests
from time import sleep
import requests
from openpyxl import Workbook,load_workbook


#kirala izle içerikler için bekletiyorsun
wp = load_workbook('C:/Users/HP/Desktop/playground/PlanSave/plansave.xlsx')
ws = wp.active



for i in ws['A']:

    #print(i.value)

    url = "http://10.98.228.146:8090/planprofiles/"+'{}'.format(i.value)

    #print(url)

    payload={}
    headers = {}

    #print(payload)

    response = requests.request("GET", url, headers=headers, data=payload)

    print(i.value ,'labelının bilgileri alındı')
    #print(response.text)

    #json=response.json()



    #print()
   # print("####################################################################################################")


    # #print(jso)
    print("####################################################################################################")


    url1 = "http://10.98.228.146:8090/planprofiles/"
    payload1=response
    headers1 = {
      'Content-Type': 'application/json'
    }


    response1 = requests.request("PUT", url1, headers=headers1, data=payload1)

    print(response1.text)
