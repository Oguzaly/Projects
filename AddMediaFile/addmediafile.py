from openpyxl import Workbook,load_workbook
import requests

wp = load_workbook("C:/Users/HP/Desktop/playground/AddMediaFile/content_id.xlsx")
ws = wp.active


# for i in  ws['A']:
#
#     url = "http://10.98.228.146:8090/contents/"+'{}'.format(i.value)+"/mediafiles/add"
#     payload="{\r\n    \"contentMediaFiles\": [\r\n        {\r\n            \"fileName\": \"test25thHour2.mp4\",\r\n            \"fileSize\": 22,\r\n            \"mediaType\": \"RawVideo\",\r\n            \"videoFormat\": \"HD\"\r\n        }\r\n    ],\r\n    \"thirdPartyTasks\": \"\"\r\n}"
#     headers = {
#       'Content-Type': 'application/json'
#     }
#     response = requests.request("POST", url, headers=headers, data=payload)
#     print(response.text)
#
# print('Video Add işlemi bitti')
#
# for i in ws['A']:
#
#     import requests
#     url = "http://10.98.228.146:8090/contents/"+'{}'.format(i.value)+"/censorshippreview"
#     payload="{\"contentAdviceShort\":\"[]\",\"status\":\"NoNeed\"}"
#     headers = {
#       'Content-Type': 'application/json'
#     }
#     response = requests.request("POST", url, headers=headers, data=payload)
#     print(response.text)
#
# print('Cencorship Adımı Bitti')
#
# for i in ws['A']:
#
#     import requests
#     url = "http://10.98.228.146:8090/contents/"+'{}'.format(i.value)+"/videooperationprocesses/Encode/Pass"
#     payload="{\"comment\":\"Operation passed by operator\"}"
#     headers = {
#       'Content-Type': 'application/json'
#     }
#     response = requests.request("POST", url, headers=headers, data=payload)
#     print(response.text)
#
# print('Encode Pass adımı tamamlandı')
# 
# for i in ws['A']:
#
#     import requests
#     url = "http://10.98.228.146:8090/contents/"+'{}'.format(i.value)+"/videooperationprocesses/Transcoding1st/direct"
#     payload="{\"processData\":null}"
#     headers = {
#       'Content-Type': 'application/json'
#     }
#     response = requests.request("POST", url, headers=headers, data=payload)
#     print(response.text)
#
# print('Direct Adımı Tamamlandı')

for i in ws['A']:

    import requests
    url = "http://10.98.228.146:8090/contents/"+'{}'.format(i.value)+"/videooperationprocesses/Transcoding1st/Start"
    payload="{\"processData\":\"{}\",\"comment\":\"TranscodingOneSt is started\"}"
    headers = {
      'Content-Type': 'application/json'
    }
    response = requests.request("POST", url, headers=headers, data=payload)
    print(response.text)

print('Transcoding1st Başladı')
