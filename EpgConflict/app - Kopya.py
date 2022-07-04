import requests
import psycopg2
from openpyxl import Workbook,load_workbook


db_host = '10.98.228.153'
db_name = 'atlas_db'
db_user = 'saatcms'
db_pass = 'cms123'
port = '5432'



conn=psycopg2.connect(dbname=db_name,user=db_user,password=db_pass,host=db_host,port=port)
cur=conn.cursor()


wp = load_workbook('C:/Users/HP/Desktop/playground/EpgConflict/ChannelId/ChannelID.xlsx')
ws = wp.active


for a in ws['A']:


    cur.execute("select p.epg_program_id , c.name , p.release_date ,to_timestamp(p.start_time/1000)::time , to_timestamp(p.end_time/1000)::time  from atlas_cms_btv.program p left join atlas_cms_btv.program p2 on p2.epg_program_id =p.epg_program_id left join atlas_cms_btv.channel c on c.id =p.channel_id where p.release_date > (current_date-7) and c.id ="+'{}'.format(a.value)+"order by p.release_date,to_timestamp(p.start_time/1000)::time")
    frsh =cur.fetchall()
    print(len(frsh))

    try:

        with open('C:/Users/HP/Desktop/playground/EpgConflict/Conflict.csv','a') as f:
            for i in range(0,len(frsh)):

                if frsh[i][4] == frsh[i+1][3]:
                    pass
                    print(i,frsh[i][4],frsh[i+1][3])
                    print(i)

                else:
                    print(i,frsh[i][0],frsh[i][2],frsh[i][4],frsh[i+1][3],'Conflict')
                    f.write('{}'.format(frsh[i][0])+','+'{}'.format(frsh[i][1])+','+'{}'.format(frsh[i][2])+','+'{}'.format(frsh[i][3])+','+'{}'.format(frsh[i][4])+'\n')

    except Exception as e:
        print('Bitti')

