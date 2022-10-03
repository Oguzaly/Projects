import requests
import json
from openpyxl import Workbook,load_workbook


wp = load_workbook('C:/Users/HP/Desktop/playground/postereksikleri/Columncode.xlsx')
ws = wp.active

print (len(ws['A']))
for i in range(1,len(ws['A'])):
#print(ws['A{}'.format(i)].value,ws['B{}'.format(i)].value)
    url = "https://itvepg10013.tmp.tivibu.com.tr/iptvepg/frame3046/sdk_getvodlist.jsp?columncode="+'{}'.format(ws['A{}'.format(i)].value)+"&columntype="+'{}'.format(ws['B{}'.format(i)].value)+"&seriestype=1,3&pageno=1&numperpage=500&ordertype=5"
    #print(url)

    payload={}
    headers = {
      'Cookie': 'JSESSIONID=53E64E57F1D8ECAC96EE4943D5C0B696; TS011688e6=012845365a3376ec51d60ca9ac83fe366fd10ef1aacb9d7c0efdfa868813ae9f7a800572c3022ba787e72d17d07e3dd19dafeb8cb6'
    }
    response = requests.request("GET", url, headers=headers, data=payload)
    #print(response.text)
    #print(type(response))
    json_data = json.loads(response.text)
    for attr  in json_data['data']:
        #print(len(attr['posterfilelist']),attr['contentcode'])
        if (len(attr['posterfilelist']) != 183):
            print(len(attr['posterfilelist']),'->',attr['contentcode'],'->',attr['genre'],'->',attr['programname'])

    #
    #
    # print(json_data[5].keys())
    # vlas = get_value(json_data,'posterfilelist')[0]
    # print(vlas)
    # print(json.dumps(json_data, indent = 1, sort_keys=True))
