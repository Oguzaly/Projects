import psycopg2
from time import sleep
import requests
from openpyxl import Workbook,load_workbook
import re
from time import sleep



wp = load_workbook('C:/Users/HP/Desktop/playground/PlanDelete/plandel.xlsx')
ws = wp.active


for i in ws['A']:

        url1 = "http://10.98.225.178:8090/planprofiles/"+'{}'.format(i.value)
        payload1={}
        headers1 = {}
        response1 = requests.request("DELETE", url1, headers=headers1, data=payload1)
        print(response1.text)
        print(i.value,' no lu plan silindi')
