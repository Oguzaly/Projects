import psycopg2
from time import sleep
import requests
from openpyxl import Workbook,load_workbook

#kirala izle içerikler için bekletiyorsun
wp = load_workbook('C:/Users/HP/Desktop//playground/epgprogramingest/epg.xlsx')
ws = wp.active


for row in ws['A']:



    url = "http://10.98.225.178:8091/programs/"+'{}'.format(row.value)+"/ingest"
    payload={}
    headers = {}

    print(row.value , ' id li epg ye ingest atıldı')


#    print(url)
#    print(i)

    response = requests.request("POST", url, headers=headers, data=payload)
    print(response.text)
