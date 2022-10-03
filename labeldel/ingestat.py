import psycopg2
from time import sleep
import requests
from openpyxl import Workbook,load_workbook

#kirala izle içerikler için bekletiyorsun
wp = load_workbook('C:/Users/HP/Desktop/playground/labeldel/labeldel.xlsx')
ws = wp.active


for row in ws['A']:



    url = "http://10.98.228.146:8090/labels/"+'{}'.format(row.value)
    payload={}
    headers = {}

    print(row.value , ' id li label silindi')


#    print(url)
#    print(i)

    response = requests.request("DELETE", url, headers=headers, data=payload)
    print(response.text)
