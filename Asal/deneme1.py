import openpyxl
from openpyxl import Workbook,load_workbook
import xlsxwriter

wb = load_workbook("C:\\Users\\HP\\Desktop\\Asal\\asal.xlsx") 
ws = wb.active

rowSayi=1
rowHarf=1
rowWrite=0

for i in range(29):
    rowSayi=rowSayi+1
    sayi=ws.cell(row=rowSayi, column=1).value
    rowHarf=rowHarf+1
    harf = ws.cell(row=rowHarf, column=2).value
    print('Sayı :'+str(sayi)+' | Harf :'+str(harf))
    print(i)




    for i in range(2, sayi):
        if (sayi % i) == 0:
            break
    else:
        print('sayi :'+str(sayi) +' , Harf:'+ str(harf))






