import requests
from openpyxl import Workbook,load_workbook


wp = load_workbook('C:/Users/HP/Desktop/playground/WaitingForAuditing/guid.xlsx')
ws = wp.active

with open('C:/Users/HP/Desktop/playground/WaitingForAuditing/status.csv','a') as f:

    for row in ws['A']:


        url = "http://10.98.225.184:8084/mw/api/getcontentinfo?id="+'{}'.format(row.value)+"&type=1"

        payload={}
        headers = {}

        response = requests.request("GET", url, headers=headers, data=payload)

        print(response.text)
        print(i)
        f.write(response.text+'\n')
